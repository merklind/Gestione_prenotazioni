from datetime import timedelta, date

from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.styles.borders import BORDER_MEDIUM, BORDER_NONE
from openpyxl.styles.colors import RED
from openpyxl.styles.fills import PatternFill, FILL_NONE
from openpyxl.utils import get_column_letter



booking_fill = PatternFill(start_color='FFEC3223', end_color='FFEC3223', fill_type='solid')
expedia_fill = PatternFill(start_color='FFF39C38', end_color='FFF39C38', fill_type='solid')
airbnb_fill = PatternFill(start_color='FFF7C143', end_color='FFF7C143', fill_type='solid')
direct_fill = PatternFill(start_color='FFFFFD54', end_color='FFFFFD54', fill_type='solid')
no_show_fill = PatternFill(start_color='FF8797AE', end_color='FF8797AE', fill_type='solid')
penalty_fill = PatternFill(start_color='FFF092D4', end_color='FFF092D4', fill_type='solid')
blank_fill = PatternFill(patternType=FILL_NONE)
non_refundable_fill = PatternFill(start_color='FF4CAC5B', end_color='FF4CAC5B', fill_type='solid')
refundable_fill = PatternFill(start_color='FF48AFEA', end_color='FF48AFEA', fill_type='solid')
confirmed_reservation_fill = PatternFill(start_color='FF4BAC5B', end_color='FF4BAC5B', fill_type='solid')
canceled_reservation_fill = PatternFill(start_color='FFEC3323', end_color='FFEC3323', fill_type='solid')
no_show_reservation_fill = PatternFill(start_color='FFF39C38', end_color='FFF39C38', fill_type='solid')



def fill_first_cell(current_cell, channel, state=''):
    if channel.lower() == 'booking':
        current_cell.fill = booking_fill
    elif channel.lower() == 'airbnb':
        current_cell.fill = airbnb_fill
    elif channel.lower() == 'expedia':
        current_cell.fill = expedia_fill
    elif channel.lower() in ('diretto', 'prolungamento', 'anticipo arrivo', 'kyos', 'ritorno', 'saltato booking'\
                               'saltato sito', 'sito sahi be', 'subito', 'saltato airbnb', 'saltato expedia'):
        current_cell.fill = direct_fill
    if state.lower() == 'no show':
        current_cell.fill = no_show_fill


def fill_cell(rate, current_cell, state=''):
    if rate.lower() == 'non rimborsabile':
        current_cell.fill = non_refundable_fill
    elif rate.lower() == 'rimborsabile':
        current_cell.fill = refundable_fill
    if state.lower() == 'no show':
        current_cell.fill = no_show_fill
    elif state.lower() == 'penale':
        current_cell.fill =penalty_fill


def start_reservation_border():
    upper_side_border = Border(top=Side(border_style=BORDER_MEDIUM), left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM))
    return upper_side_border


def middle_reservation_border():
    side_border = Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM))
    return side_border


def end_reservation_border():
    bottom_border = Border(bottom=Side(border_style=BORDER_MEDIUM), left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM))
    return bottom_border


def single_day_reservation_border():
    total_border = Border(
        left=Side(border_style=BORDER_MEDIUM),
        right=Side(border_style=BORDER_MEDIUM),
        top=Side(border_style=BORDER_MEDIUM),
        bottom=Side(border_style=BORDER_MEDIUM)
    )
    return total_border


def set_value_gross_cell(current_cell, current_date_cell, price_per_night, rate):
    if current_date_cell.value is not None:
        if current_cell.value is None:
            current_cell.value = price_per_night
        elif current_cell.value is not None:
            old_price = current_cell.value
            current_cell.value = f'={old_price}+{price_per_night}'
        fill_cell(rate, current_cell)


def set_value_occupation_cell(current_cell):
    if current_cell.value is None:
        current_cell.value = 1
    elif current_cell.value is not None:
        old_value = current_cell.value
        current_cell.value = old_value + 1


def set_rendiconto_occupation_cell(ws_rendiconto, start_row_rendiconto, info_reservation, occupation_column):

    date_check_in = info_reservation['Entrata']
    date_check_out = info_reservation['Uscita']
    channel = info_reservation['Tramite']
    state = info_reservation['Stato prenotazione']
    rate = info_reservation['Tipo tariffa']

    offset_date = 0
    current_cell = ws_rendiconto.cell(row=start_row_rendiconto + offset_date, column=occupation_column)

    # se la prenotazione è per più di una notte
    if date_check_in != date_check_out - timedelta(days=1):
        current_cell.border = start_reservation_border()
        set_value_occupation_cell(current_cell)
        fill_first_cell(current_cell, channel, state)
        offset_date += 1

        current_date_cell = ws_rendiconto.cell(row=start_row_rendiconto + offset_date, column=1)
        current_cell = ws_rendiconto.cell(row=start_row_rendiconto + offset_date, column=occupation_column)

        while current_date_cell.value != date_check_out:
            if current_date_cell.value is not None and date_check_in < current_date_cell.value < date_check_out - timedelta(days=1):
                current_cell.border = middle_reservation_border()
                set_value_occupation_cell(current_cell)
                fill_cell(rate, current_cell, state)

            elif current_date_cell.value == date_check_out - timedelta(days=1):
                current_cell.border = end_reservation_border()
                set_value_occupation_cell(current_cell)
                fill_cell(rate, current_cell, state)

            offset_date += 1
            current_date_cell = ws_rendiconto.cell(row=start_row_rendiconto + offset_date, column=1)
            current_cell = ws_rendiconto.cell(row=start_row_rendiconto + offset_date, column=occupation_column)

    # se la prenotazione è solo per una notte
    if date_check_in == date_check_out - timedelta(days=1):
        current_cell.border = single_day_reservation_border()
        set_value_occupation_cell(current_cell)
        fill_first_cell(current_cell, channel)


def set_value_net_cell(ws_rendiconto, current_date_cell, current_cell, channel, gross_column, start_row_rendiconto, offset_date):
    column_letter = get_column_letter(gross_column)
    row_index = start_row_rendiconto + offset_date

    if current_date_cell.value is not None:

        if channel.lower() in ('booking', 'expedia'):

            if current_cell.value is None:
                current_cell.value = "=" + column_letter + str(row_index) + "*0.82"

            elif current_cell.value is not None:
                result = ''
                old_gross_price = ws_rendiconto.cell(row_index, gross_column).value.split('+')[0][1:]
                new_gross_price = ws_rendiconto.cell(row_index, gross_column).value.split('+')[1]
                type_multiply = current_cell.value.split('*')[1:]
                for multiplier in type_multiply:
                    result = result+multiplier+'*'
                if len(type_multiply) != 0:
                    current_cell.value = f'={old_gross_price}*{result[:-1]} + {new_gross_price}*0.82'
                else:
                    current_cell.value = f'={old_gross_price} + {new_gross_price}*0.82'

        if channel.lower() in ('diretto', 'prolungamento', 'anticipo arrivo', 'kyos', 'ritorno', 'saltato booking'\
                               'saltato sito', 'sito sahi be', 'subito', 'saltato airbnb', 'saltato expedia'):
            if current_cell.value is None:
                current_cell.value = '=' + column_letter + str(row_index)
            elif current_cell.value is not None:
                result = ''
                old_gross_price = ws_rendiconto.cell(row_index, gross_column).value.split('+')[0][1:]
                new_gross_price = ws_rendiconto.cell(row_index, gross_column).value.split('+')[1]
                type_multiply = current_cell.value.split('*')[1:]
                for multiplier in type_multiply:
                    result = result + multiplier + '*'
                if len(type_multiply) != 0:
                    current_cell.value = f'={old_gross_price}*{result[:-1]} + {new_gross_price}'
                else:
                    current_cell.value = f'={old_gross_price} + {new_gross_price}'

        if channel.lower() in ('airbnb'):
            if  date(2018, 1, 3) <= current_date_cell.value.date() <= date(2018, 7, 14):
                if current_cell.value is None:
                    current_cell.value = f'={column_letter}{row_index} * (1 - 3%*1.22)'
                elif current_cell.value is not None:
                    result = ''
                    old_gross_price = ws_rendiconto.cell(row_index, gross_column).value.split('+')[0][1:]
                    new_gross_price = ws_rendiconto.cell(row_index, gross_column).value.split('+')[1]
                    type_multiply = current_cell.value.split('*')[1:]
                    for multiplier in type_multiply:
                        result = result + multiplier + '*'
                    if len(type_multiply) != 0:
                        current_cell.value = f'={old_gross_price}*{result[:-1]} + {new_gross_price}*(1-3%*1.22)'
                    else:
                        current_cell.value = f'={old_gross_price} + {new_gross_price}*(1-3%*1.22)'
            elif date(2018, 12, 18) <= current_date_cell.value.date() <= date(2019, 5, 19):
                if current_cell.value is None:
                    current_cell.value = f'={column_letter}{row_index} * (1 - 5%*1.22)'
                elif current_cell is not None:
                    result = ''
                    old_gross_price = ws_rendiconto.cell(row_index, gross_column).value.split('+')[0][1:]
                    new_gross_price = ws_rendiconto.cell(row_index, gross_column).value.split('+')[1]
                    type_multiply = current_cell.value.split('*')[1:]
                    for multiplier in type_multiply:
                        result = result + multiplier + '*'
                    if len(type_multiply) != 0:
                        current_cell.value = f'={old_gross_price}*{result[:-1]} + {new_gross_price}*(1-5%*1.22)'
                    else:
                        current_cell.value = f'={old_gross_price} + {new_gross_price}*(1-5%*1.22)'
            elif date(2020, 2, 17) <= current_date_cell.value.date():
                if current_cell.value is None:
                    current_cell.value = f'={column_letter}{row_index} * (1 - 14%*1.22)'
                elif current_cell.value is not None:
                    result = ''
                    old_gross_price = ws_rendiconto.cell(row_index, gross_column).value.split('+')[0][1:]
                    new_gross_price = ws_rendiconto.cell(row_index, gross_column).value.split('+')[1]
                    type_multiply = current_cell.value.split('*')[1:]
                    for multiplier in type_multiply:
                        result = result + multiplier + '*'
                    if len(type_multiply) != 0:
                        current_cell.value = f'={old_gross_price}*{result[:-1]} + {new_gross_price}*(1-14%*1.22)'
                    else:
                        current_cell.value = f'={old_gross_price} + {new_gross_price}*(1-14%*1.22)'



def set_rendiconto_gross_cell(ws_rendiconto, ws_riepilogo, gross_column, price_column, start_row_rendiconto,
                              reservation, info_reservation):

    date_check_in = info_reservation['Entrata']
    date_check_out = info_reservation['Uscita']
    price_per_night = round(info_reservation['Importo notte'], 2)
    nights = info_reservation['Giorni']
    rate = info_reservation['Tipo tariffa']

    offset_date = 0
    offset_price_per_night = 0

    current_cell_price = ws_riepilogo.cell(row=reservation, column=price_column + offset_price_per_night)
    current_date_cell = ws_rendiconto.cell(row=start_row_rendiconto, column=1)
    current_cell = ws_rendiconto.cell(row=start_row_rendiconto, column=gross_column)

    # se la prenotazione è per più di una notte
    if date_check_in != date_check_out - timedelta(days=1):
        # se la prenotazione è maggiore di 31 giorni
        if nights > 31:
            # per ogni notte imposto il prezzo medio per notte
            while current_date_cell.value != date_check_out:
                set_value_gross_cell(current_cell, current_date_cell, price_per_night, rate)
                offset_date += 1
                current_cell = ws_rendiconto.cell(row=start_row_rendiconto + offset_date, column=gross_column)
                current_date_cell = ws_rendiconto.cell(row=start_row_rendiconto + offset_date, column=1)
        # se la prenotazione è minore o uguale a 31 giorni
        elif nights <= 31:
            # per ogni notte imposta il prezzo a notte reale
            while current_date_cell.value != date_check_out:
                price_per_night = current_cell_price.value
                set_value_gross_cell(current_cell, current_date_cell, price_per_night, rate)
                if current_date_cell.value is not None:
                    offset_price_per_night += 1
                offset_date += 1

                current_cell_price = ws_riepilogo.cell(row=reservation, column=price_column + offset_price_per_night)
                current_cell = ws_rendiconto.cell(row=start_row_rendiconto + offset_date, column=gross_column)
                current_date_cell = ws_rendiconto.cell(row=start_row_rendiconto + offset_date, column=1)

    # se la prenotazione è di una sola notte
    elif date_check_in == date_check_out - timedelta(days=1):
        set_value_gross_cell(current_cell, current_date_cell, price_per_night, rate)
        fill_cell(rate, current_cell)


def set_rendiconto_net_cell(ws_rendiconto, start_row_rendiconto, net_column, gross_column, info_reservation):

    date_check_out = info_reservation['Uscita']
    channel = info_reservation['Tramite']

    offset_date = 0
    current_date_cell = ws_rendiconto.cell(row=start_row_rendiconto + offset_date, column=1)
    current_cell = ws_rendiconto.cell(row=start_row_rendiconto + offset_date, column=net_column)

    # imposto il prezzo netto per ogni notte
    while current_date_cell.value != date_check_out:
        set_value_net_cell(ws_rendiconto, current_date_cell, current_cell, channel, gross_column, start_row_rendiconto, offset_date)
        offset_date += 1
        current_date_cell = ws_rendiconto.cell(row=start_row_rendiconto + offset_date, column=1)
        current_cell = ws_rendiconto.cell(row=start_row_rendiconto + offset_date, column=net_column)


def set_cell_reservation_break(ws_rendiconto, ws_rendiconto_data_only, start_row_rendiconto, info_reservation, taken_column, gross_column,
                               net_column, today, rate):

    offset_date = 0
    current_date_cell = ws_rendiconto.cell(row=start_row_rendiconto + offset_date, column=1)
    while current_date_cell.value is None or current_date_cell.value != today:
        offset_date += 1
        current_date_cell = ws_rendiconto.cell(row=start_row_rendiconto + offset_date, column=1)

    days_remaining = (info_reservation['Uscita'] - today).days

    while days_remaining > 0:
        if current_date_cell.value is not None:
            occupation_cell = ws_rendiconto.cell(row=start_row_rendiconto + offset_date, column=taken_column)
            set_value_occupation_cell(occupation_cell)
            fill_cell(rate, occupation_cell, info_reservation['Stato prenotazione'])
            if current_date_cell.value == info_reservation['Uscita'] - timedelta(days=1):
                occupation_cell.border = end_reservation_border()
            elif current_date_cell.value < info_reservation['Uscita'] - timedelta(days=1):
                occupation_cell.border = middle_reservation_border()
            gross_cell = ws_rendiconto.cell(row=start_row_rendiconto + offset_date, column=gross_column)
            net_cell = ws_rendiconto.cell(row=start_row_rendiconto + offset_date, column=net_column)
            if info_reservation['Giorni'] <= 31:
                set_value_gross_cell(gross_cell, current_date_cell, info_reservation['Giorno ' + str(info_reservation['Giorni'] - days_remaining + 1)], info_reservation['Tipo tariffa'])
            elif info_reservation['Giorni'] > 31:
                set_value_gross_cell(gross_cell, current_date_cell, info_reservation['Importo notte'],info_reservation['Tipo tariffa'])
            set_value_net_cell(ws_rendiconto_data_only, current_date_cell, net_cell, info_reservation['Tramite'], gross_column, start_row_rendiconto, offset_date)
            days_remaining -= 1
            offset_date += 1
            current_date_cell = ws_rendiconto.cell(row=start_row_rendiconto + offset_date, column=1)
        else:
            offset_date += 1
            current_date_cell = ws_rendiconto.cell(row=start_row_rendiconto + offset_date, column=1)


def erase_rendiconto_cell(ws_rendiconto, date_check_out, current_row, taken_column, gross_column, net_column):

    current_cell = ws_rendiconto.cell(row=current_row, column=1)
    while ws_rendiconto.cell(row=current_row, column=1).value != date_check_out:
        if current_cell.value is not None:
            current_cell = ws_rendiconto.cell(row=current_row, column=taken_column)
            erase_cell(current_cell)
            current_cell = ws_rendiconto.cell(row=current_row, column=gross_column)
            erase_cell(current_cell)
            current_cell = ws_rendiconto.cell(row=current_row, column=net_column)
            erase_cell(current_cell)
            current_row += 1
        else:
            current_row += 1


def erase_cell(current_cell):
    current_cell.fill = PatternFill(patternType=FILL_NONE)
    current_cell.border = BORDER_NONE
    current_cell.value = None


def erase_all_reservation_rendiconto(ws_rendiconto, start_row_rendiconto, end_row_rendiconto):

    offset_row = 0
    for index in range(start_row_rendiconto, end_row_rendiconto + 1):
        current_cell = ws_rendiconto.cell(row=start_row_rendiconto + offset_row, column=1)
        if current_cell.value is not None:
            for cell in ws_rendiconto.iter_cols(min_col=3, max_col=75, min_row=start_row_rendiconto + offset_row, max_row=start_row_rendiconto + offset_row):
                erase_cell(cell[0])
        offset_row += 1
    print('Erased all info in Rendiconto worksheet\n')


def erase_future_reservation(ws_rendiconto, today, start_row_rendiconto, end_row_rendiconto):
    '''
    from ws_rendiconto erase all content's cell from today till the end of calendar
    :param ws_rendiconto: worksheet object
    :param today: the date of today
    :param start_row_rendiconto: number of row from which start
    :param end_row_rendiconto: number of row in which finish
    '''
    # initialize some variables
    offset_date = 0
    current_date_cell = ws_rendiconto.cell(row=start_row_rendiconto + offset_date, column=1).value

    # find today's row
    while current_date_cell is None or current_date_cell < today:
        offset_date += 1
        current_date_cell = ws_rendiconto.cell(row=start_row_rendiconto + offset_date, column=1).value


    print('Prima riga da cancellare: ' + str(start_row_rendiconto+offset_date))

    # loop from starting row till the end of Rendiconto worksheet
    for index in range(start_row_rendiconto, end_row_rendiconto + 1):
        # get the current date cell
        current_cell = ws_rendiconto.cell(row=start_row_rendiconto + offset_date, column=1)
        # if the current date cell is not blank
        if current_cell.value is not None:
            # loop over the column of all apartments
            for cell in ws_rendiconto.iter_cols(min_col=3, max_col=75, min_row=start_row_rendiconto + offset_date,
                                                max_row=start_row_rendiconto + offset_date):
                # delete content of the cell
                erase_cell(cell[0])
        # increase the offset for the current date cell
        offset_date += 1
    print('Erased all info in Rendiconto worksheet\n')


def stylish_apartment_cell(ws, reservation, starting_row, column_price):
    apartment_res = reservation['apartment']
    ws.merge_cells(start_row=starting_row, end_row=starting_row, start_column=column_price, end_column=column_price + 2)
    current_cell = ws.cell(row=starting_row, column=column_price)
    current_cell.value = apartment_res
    current_cell.alignment = Alignment(horizontal='center')

    starting_row += 1
    return starting_row


def stylish_name_guest_cell(ws, reservation, starting_row, column_price):

    rate = reservation['rate'].lower()
    state = reservation['state'].lower()
    name_guest = reservation['name_guest']

    ws.merge_cells(start_row=starting_row, end_row=starting_row, start_column=column_price, end_column=column_price + 2)
    current_cell = ws.cell(row=starting_row, column=column_price)
    current_cell.value = name_guest
    current_cell.alignment = Alignment(horizontal='center')
    if state == 'cancellata':
        current_cell.font = Font(b=True, color=RED)
    fill_cell(rate, current_cell, state)

    starting_row += 1

    return starting_row


def stylish_header_cell(ws, starting_row, column_price):

    ws.cell(row=starting_row, column=column_price).value = 'Data'
    ws.cell(row=starting_row, column=column_price).alignment = Alignment(horizontal='center')
    ws.cell(row=starting_row, column=column_price + 1).value = 'Importo'
    ws.cell(row=starting_row, column=column_price + 1).alignment = Alignment(horizontal='center')
    ws.cell(row=starting_row, column=column_price + 2).value = 'Tramite'
    ws.cell(row=starting_row, column=column_price + 2).alignment = Alignment(horizontal='center')

    starting_row += 1

    return starting_row


def stylish_channel_cell(ws, reservation, starting_row, column_price):

    channel = reservation['channel']
    current_cell = ws.cell(row=starting_row, column=column_price + 2)
    current_cell.value = channel
    current_cell.font = Font(color=RED)

    return starting_row


def stylish_price_per_night(ws, reservation, starting_row, column_price):

    days = reservation['days']
    check_in_year = reservation['check-in'].year
    check_in_month = reservation['check-in'].month
    check_in_day = reservation['check-in'].day
    date_check_in = date(check_in_year, check_in_month, check_in_day)
    if days <= 31:
        for offset_date in range(days):
            current_date_cell = ws.cell(row=starting_row + offset_date, column=column_price)
            current_price_cell = ws.cell(row=starting_row + offset_date, column=column_price + 1)

            current_date_cell.value = date_check_in + timedelta(days=offset_date)
            current_price_cell.value = reservation[f'price{offset_date+1}']

            current_date_cell.number_format = '[$-F800]dddd\\,\\ mmmm\\ dd\\,\\ yyyy'
            current_price_cell.style = 'Migliaia'

    elif days > 31:
        check_out_year = reservation['check-out'].year
        check_out_month = reservation['check-out'].month
        check_out_day = reservation['check-out'].day
        date_check_out = date(check_out_year, check_out_month, check_out_day)
        total_price = reservation['total_price']
        for offset_date in range(2):
            current_date_cell = ws.cell(row=starting_row + offset_date, column=column_price)
            current_price_cell = ws.cell(row=starting_row + offset_date, column=column_price + 1)
            if offset_date == 0:
                current_date_cell.value = date_check_in
            elif offset_date == 1:
                current_date_cell.value = date_check_out

            current_price_cell.value = round(total_price/2, ndigits=2)
            current_price_cell.style = 'Migliaia'

            current_date_cell.number_format = '[$-F800]dddd\\,\\ mmmm\\ dd\\,\\ yyyy'
            current_price_cell.style = 'Migliaia'

    starting_row += (offset_date+1)

    return starting_row


def stylish_sum_price(ws, reservation, starting_row, column_price):

    days = reservation['days']
    letter_column = get_column_letter(column_price + 1)

    current_cell = ws.cell(row=starting_row, column=column_price)
    current_cell.value = 'Totale'
    current_cell.alignment = Alignment(horizontal='right')
    current_cell.font = Font(color=RED)

    current_cell = ws.cell(row=starting_row, column=column_price+1)
    if days > 31:
        current_cell.value = f'=SUM({letter_column}{starting_row-2}:{letter_column}{starting_row-1})'


    elif days <= 31:
        current_cell.value = f'=SUM({letter_column}{starting_row-days}:{letter_column}{starting_row-1})'

    current_cell.style = 'Migliaia'

    starting_row += 2

    return starting_row