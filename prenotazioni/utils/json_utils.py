from json import dump, load
from pathlib import PurePath

from openpyxl import load_workbook

from prenotazioni.constants import *
from prenotazioni.utils.path_utils import get_folder_path

from prenotazioni.constants import MAX_ROW_FILE


def find_column_apartment(info_reservation, json_apartment, type_column):

    house = info_reservation['Residenza'].lower()
    apartment = info_reservation['App. assegnato'].lower()
    column = json_apartment[house][type_column][apartment]
    return column


def generate_json(ws_file, column_file):

    folder_path = get_folder_path()
    file_path = PurePath.joinpath(folder_path, ws_file)
    column_path = PurePath.joinpath(folder_path, column_file)
    wb = load_workbook(file_path, data_only=True)
    ws_riepilogo = wb[RIEPILOGO_WS]
    column_label = {}

    column_apartment_file = open(file=str(column_path), mode='w')
    for current_column in range(1, 91):
        current_cell = ws_riepilogo.cell(row=1, column=current_column)
        if current_cell.value is not None:
            column_label[current_cell.value] = current_column

        current_column += 1

    dump(column_label, column_apartment_file, indent=4, ensure_ascii=False)
    print('JSON generato')


def reset_max_row_json():
    folder_path = get_folder_path()

    max_row_json = open(str(folder_path.joinpath(MAX_ROW_FILE)))
    max_row_dettaglio_prezzo = load(max_row_json)
    max_row_json.close()

    for year in max_row_dettaglio_prezzo.keys():
        for house in max_row_dettaglio_prezzo[year]:
            max_row_dettaglio_prezzo[year][house] = 2

    max_row_json = open(str(folder_path.joinpath(MAX_ROW_FILE)), mode='w')
    dump(max_row_dettaglio_prezzo, max_row_json, indent=4)
    max_row_json.close()


def save_max_row_json(max_row_dettaglio_prezzo):

    max_row_dettaglio_prezzo_json = open('../resource/max_row_dettaglio_prezzo.json', mode='w+')
    dump(max_row_dettaglio_prezzo, max_row_dettaglio_prezzo_json, indent=4)
