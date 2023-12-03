from copy import deepcopy
from datetime import date, timedelta
from json import load
from pathlib import Path, PurePath

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, colors
from openpyxl.utils import get_column_letter

from prenotazioni.style_worksheet.style_excel import no_show_reservation_fill, confirmed_reservation_fill, \
    canceled_reservation_fill, stylish_apartment_cell, stylish_name_guest_cell, stylish_header_cell, \
    stylish_channel_cell, stylish_price_per_night, stylish_sum_price


def find_max_row(ws):
    '''
    Find the max row in the worksheet. The first row in the first column that is None.
    :param ws: the worksheet where to find the number of the max row
    :return: the number of the max row.
    '''
    row = 1
    current_cell = ws.cell(row=row, column=1).value
    while current_cell is not None:
        row += 1
        current_cell = ws.cell(row=row, column=1).value

    return row


def get_info_reservation(ws_riepilogo_data_only, number_reservation, column_label):
    info_reservation = {}

    for entry in column_label.items():
        info_reservation[entry[0]] = ws_riepilogo_data_only.cell(row=number_reservation, column=entry[1]).value

    return info_reservation


def insert_new_rendiconto_reservation(ws_dettaglio_prezzo, info_reservation):
    current_path = Path(__file__).parent.absolute()

    json_file = open(str(PurePath(current_path).joinpath('resource', 'apartment.json')))
    apartment_file = load(json_file)
    info_reservation['start_column'] = apartment_file[info_reservation['house']]['Dettaglio prezzi']
    row = 1
    offset_day = 0

    current_cell = ws_dettaglio_prezzo.cell(row=row, column=info_reservation['start_column'])
    row += 1
    next_cell = ws_dettaglio_prezzo.cell(row=row, column=info_reservation['start_column'])

    if info_reservation['nights'] <= 31:
        while current_cell.value is not None or next_cell.value is not None:
            current_cell = next_cell
            row += 1
            next_cell = ws_dettaglio_prezzo.cell(row=row, column=info_reservation['start_column'])

        ws_dettaglio_prezzo.merge_cells(start_row=row, end_row=row, start_column=info_reservation['start_column'],
                                        end_column=info_reservation['start_column'] + 2)
        ws_dettaglio_prezzo.cell(row=row, column=info_reservation['start_column']).value = info_reservation['apartment']
        ws_dettaglio_prezzo.cell(row=row, column=info_reservation['start_column']).alignment = Alignment(
            horizontal='center')

        row += 1
        ws_dettaglio_prezzo.merge_cells(start_row=row, end_row=row, start_column=info_reservation['start_column'],
                                        end_column=info_reservation['start_column'] + 2)
        ws_dettaglio_prezzo.cell(row=row, column=info_reservation['start_column']).value = info_reservation[
            'name_guest']
        ws_dettaglio_prezzo.cell(row=row, column=info_reservation['start_column']).alignment = Alignment(
            horizontal='center')
        if info_reservation['state'].lower() == 'cancellata':
            ws_dettaglio_prezzo.cell(row=row, column=info_reservation['start_column']).fill = canceled_reservation_fill
        elif info_reservation['state'].lower() == 'prenotata' or info_reservation['state'].lower() == 'confermata':
            ws_dettaglio_prezzo.cell(row=row, column=info_reservation['start_column']).fill = confirmed_reservation_fill
        elif info_reservation['state'].lower() == 'no show':
            ws_dettaglio_prezzo.cell(row=row, column=info_reservation['start_column']).fill = no_show_reservation_fill

        row += 1
        ws_dettaglio_prezzo.cell(row=row, column=info_reservation['start_column']).value = 'Data'
        ws_dettaglio_prezzo.cell(row=row, column=info_reservation['start_column']).alignment = Alignment(
            horizontal='center')
        ws_dettaglio_prezzo.cell(row=row, column=info_reservation['start_column'] + 1).value = 'Importo'
        ws_dettaglio_prezzo.cell(row=row, column=info_reservation['start_column'] + 2).value = 'Tramite'

        row += 1
        first_row = row

        ws_dettaglio_prezzo.cell(row=row, column=info_reservation['start_column'] + 2).value = info_reservation[
            'channel']

        year = info_reservation['date_check_in'].year
        month = info_reservation['date_check_in'].month
        day = info_reservation['date_check_in'].day

        for current_cell in ws_dettaglio_prezzo.iter_rows(min_row=row,
                                                          max_row=row + info_reservation['nights'] - 1,
                                                          min_col=info_reservation['start_column'],
                                                          max_col=info_reservation['start_column'] + 1):
            current_cell[0].number_format = '[$-F800]dddd\\,\\ mmmm\\ dd\\,\\ yyyy'
            current_cell[0].value = date(year, month, day) + timedelta(days=offset_day)
            current_cell[1].value = info_reservation['price']['day' + str(offset_day)]
            current_cell[1].style = 'Migliaia'
            offset_day += 1
            row += 1

        ws_dettaglio_prezzo.cell(row=row, column=info_reservation['start_column']).value = 'Totale'
        ws_dettaglio_prezzo.cell(row=row, column=info_reservation['start_column']).alignment = Alignment(
            horizontal='right')

        ws_dettaglio_prezzo.cell(row=row, column=info_reservation['start_column'] + 1).value = \
            '=SUM(' + str(get_column_letter(info_reservation['start_column'] + 1)) + str(first_row) + ':' + \
            str(get_column_letter(info_reservation['start_column'] + 1)) + str(
                first_row + info_reservation['nights'] - 1) + ')'
        ws_dettaglio_prezzo.cell(row=row, column=info_reservation['start_column'] + 1).style = 'Migliaia'
        ws_dettaglio_prezzo.cell(row=row, column=info_reservation['start_column'] + 1).font = Font(color=colors.RED,
                                                                                                   b=True)


def open_workbook(path, data_only=False):
    '''
    Open the workbook at given path
    :param path: Path of the workbook
    :param data_only: If the workbook has to be open in data_only mode or not
    :return: the workbook object and the real path of the workbook
    '''
    try:
        # try to open the workbook at the given path
        wb = load_workbook(path, data_only=data_only)
        print('Workbook caricato')
        return wb, path
    except FileNotFoundError:
        # if the workbook at the given path doesn't exist
        # ask the user to input the name of the excel file
        print(f'Foglio Excel non trovato al percorso: {path}')
        name_file = Path(input(f'Inserisci il nome del file: '))
        name_file_with_ext = name_file.with_suffix('.xlsx')
        folder_path = Path(path).parent
        file_path = folder_path.joinpath(name_file_with_ext)
        wb = load_workbook(file_path, data_only=data_only)
        print('Workbook caricato')
        print(f'Foglio Excel {name_file_with_ext} caricato')
        # return the workbook object and the new path
        return wb, file_path


def open_worksheet(wb, name):
    '''
    Try to open the worksheet with the given name of wb object
    :param wb: workbook where to open the worksheet
    :param name: name of the worksheet
    :return: worksheet object
    '''
    try:
        # try to open the worksheet
        ws = wb[name]
    except KeyError:
        # if the worksheet with given name doesn't exits
        # create a worksheet with the given name
        ws = wb.create_sheet(title=f'{name}')
        duplicate_worksheet(wb, name)
        print(f'Il foglio non è presente.\nCreato foglio con nome {name}.')
    # return the worksheet object
    return ws


def duplicate_worksheet(wb, name_ws):
    new_ws = wb.copy_worksheet(wb['Sample'])
    wb.remove(wb[f'{name_ws}'])
    new_ws.title = f'{name_ws}'

    return new_ws


def get_reservation_by_year(ws, year, max_row, column_label):
    check_in_column = column_label['Entrata']
    all_reservation_by_year = []
    new_reservation = {}

    for row in range(2, max_row):
        current_date_check_in = ws.cell(row=row, column=check_in_column).value
        year_current_date = current_date_check_in.year
        if year_current_date == year:
            new_reservation['name_guest'] = ws.cell(row=row, column=column_label['Rif. Inq']).value
            new_reservation['house'] = ws.cell(row=row, column=column_label['Residenza']).value
            new_reservation['apartment'] = ws.cell(row=row, column=column_label['App. assegnato']).value
            new_reservation['check-in'] = ws.cell(row=row, column=column_label['Entrata']).value
            new_reservation['check-out'] = ws.cell(row=row, column=column_label['Uscita']).value
            new_reservation['days'] = ws.cell(row=row, column=column_label['Giorni']).value
            new_reservation['state'] = ws.cell(row=row, column=column_label['Stato prenotazione']).value
            new_reservation['rate'] = ws.cell(row=row, column=column_label['Tipo tariffa']).value
            new_reservation['channel'] = ws.cell(row=row, column=column_label['Tramite']).value
            if new_reservation['days'] <= 31:
                for days in range(1, new_reservation['days'] + 1):
                    new_reservation[f'price{days}'] = ws.cell(row=row, column=column_label[f'Giorno {days}']).value
            else:
                new_reservation['total_price'] = ws.cell(row=row, column=column_label[f'Importo totale']).value
            all_reservation_by_year.append((deepcopy(new_reservation)))

        new_reservation.clear()

    return all_reservation_by_year


def insert_new_dettaglio_prezzo_reservation(ws, reservation, starting_row, column_price):
    row_update = stylish_apartment_cell(ws, reservation, starting_row, column_price)
    row_update = stylish_name_guest_cell(ws, reservation, row_update, column_price)
    row_update = stylish_header_cell(ws, row_update, column_price)
    row_update = stylish_channel_cell(ws, reservation, row_update, column_price)
    row_update = stylish_price_per_night(ws, reservation, row_update, column_price)
    row_update = stylish_sum_price(ws, reservation, row_update, column_price)

    return row_update
